# EH 批量采集＆更新

import requests
import bs4
import os
import time
import shelve
from random import randint
from pprint import pprint
import copy
import openpyxl
import sys
from hashlib import sha1
from shutil import rmtree

sys.setrecursionlimit(1000000)


# Html download function
# 输入参数raw=1表示直接返回res raw=0则返回res.text
# 若下载失败， 一律返回None
def download(url, num_retries=3, cookie='', params='', raw=0, timeout=40):
    print('Downloading: ', url)
    headers = {
        'user-agent':
        'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:72.0)' +
        ' Gecko/20100101 Firefox/72.0',
        'connection':
        'keep-alive'
    }
    if cookie != '':
        headers['cookie'] = cookie
    try:
        resp = requests.get(url, headers=headers,
                            params=params, timeout=timeout)
        html = resp.text
        if resp.status_code >= 400:
            print('Download error: ', resp.text)
            html = None
            if num_retries and 500 <= resp.status_code < 600:
                return download(url, num_retries-1)
    except requests.exceptions.RequestException:
        print('Download error!!!')
        html = None
        resp = None
    except requests.exceptions.Timeout:
        print('请求超时!')
        html = None
        resp = None
    if raw == 1:
        return resp
    else:
        return html


# 获取搜索参数
def getParam():
    # 构造搜索参数
    params = {}
    params['page'] = 0          # 页数（从零开始）
    params['f_cats'] = ''       # 类别
    params['f_search'] = ''     # 搜索字段
    params['advsearch'] = 1     # 高级筛选
    params['f_sname'] = 'on'    # Search Gallery Name
    params['f_stags'] = 'on'    # Search Gallery Tags
    params['f_sr'] = 'on'       # Rating
    params['f_srdd'] = 4        # Minimum Rating
    # 可选类别
    cats = {}
    cats['Doujinshi'] = 1021
    cats['Manga'] = 1019
    cats['Artist CG'] = 1015
    cats['Game CG'] = 1007
    cats['Non-H'] = 767
    cats['Image Set'] = 991
    cats['Cosplay'] = 959
    # 获取搜索类别
    inTemp = input('搜索类别：')
    while inTemp not in cats:
        print('输入错误！')
        inTemp = input('请重新输入：')
        if inTemp == 'exit' or inTemp == 'quit':
            exit()
    params['f_cats'] = cats[inTemp]
    inTemp = ''
    # 获取搜索字段
    inTemp = input('关键词:')
    params['f_search'] = inTemp
    inTemp = ''
    # 最低评分
    inTemp = input('最低评分:')
    while int(inTemp) > 5 or int(inTemp) < 0:
        print('输入错误！')
        inTemp = input('请重新输入：')
        if inTemp == 'exit' or inTemp == 'quit':
            exit()
    params['f_srdd'] = int(inTemp)
    del inTemp
    return params


# 下载并保存搜索结果页
# 返回本地html所在的位置path
def getSearch(url, params, startPage, endPage):
    # 记录入口工作目录
    inPath = os.getcwd()
    # 创建文件夹保存搜索页
    try:
        os.makedirs('scHtml')
    except FileExistsError:
        print('scHtml folder already exists!')
    os.chdir('scHtml')
    folderPath = os.getcwd()
    # 循环下载搜索页
    pageCnt = startPage
    while pageCnt <= endPage:
        print('Deal with page ', pageCnt)
        params['page'] = pageCnt
        scHtml = download(url=url, params=params, raw=0, timeout=40)
        if scHtml is None:
            continue
        scFile = open(str(pageCnt)+'.html', 'w')
        scFile.write(scHtml)
        scFile.close()
        pageCnt += 1
    print('搜索结果页下载完成！')
    print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))
    # 重置工作目录
    os.chdir(inPath)
    return folderPath


# 从搜索结果页中提取入口url
# 保留新添加到work于newDic中
# 并更新workDic
# newDic 和　workDic 同时保存在dicFile中
# 以title的SHA1的后8为作为字典的key
# 包含title url SHA1
# 输入参数 html所在的文件夹
def getWorkUrl(htmlFolder):
    # 记录入口工作目录
    inPath = os.getcwd()
    # 将当前工作目录移入htmlFolder
    os.chdir(htmlFolder)
    # 循环处理每一页
    # 获取title 和 url 计算title的SHA1
    newDic = {}
    htmlList = os.listdir()
    for page in htmlList:
        tempDic = {}
        htmlFile = open(page, 'r')
        mainSoup = bs4.BeautifulSoup(htmlFile, 'lxml')
        htmlFile.close()
        nodeA = mainSoup.find('table', attrs={'class': 'gltc'})
        nodeB = nodeA.find_all('td', attrs={'class': 'glname'})
        for item in nodeB:
            tempDic['url'] = item.find('a').get('href')
            tempDic['title'] = item.find(
                'div', attrs={'class': 'glink'}).getText()
            workSHA1 = sha1(tempDic['title'].encode(
                encoding='UTF-8')).hexdigest()
            tempDic['SHA1'] = workSHA1
            newDic[workSHA1[-9:-1]] = copy.copy(tempDic)
    # 回到主工作目录判断dicFile是否已经存在
    os.chdir(workPath)
    tempDic = {}
    if 'dicFile' in os.listdir():
        # 更新newDic & workDic
        dicFile = shelve.open('dicFile')
        workDic = dicFile['workDic']
        dicFile.close()
        for k, v in newDic.items():
            if k not in workDic:
                tempDic[k] = v
                workDic[k] = v
        newDic = copy.copy(tempDic)
    else:
        workDic = copy.copy(newDic)
    # 保存newDic & workDic于dicFile
    dicFile = shelve.open('dicFile')
    dicFile['workDic'] = workDic
    dicFile['newDic'] = newDic
    dicFile.close()
    # 提示信息
    print('入口url 提取完毕！')
    print('待处理', len(newDic), '项')
    # 重置工作目录
    os.chdir(inPath)


# 主下载函数
# 输入参数： deWork 一件work的字典
# skip = 1 跳过缺页继续下载
# skip = 0 一旦遇到缺页 直接返回0
# 返回值：  成功返回１，　否则返回0或-1(当前work已存在)
def downWork(deWork, skip=0):
    # 记录入口工作目录
    inPath = os.getcwd()
    # 下载详情页(相当于测试网络)
    deUrl = deWork['url']
    dePage = download(url=deUrl)
    if dePage is None:
        print('Fail to download detail page!')
        return 0
    # 创建该作品的文件夹
    try:
        os.mkdir(deWork['SHA1'][-9:-1])
    except FileExistsError:
        print('Work already exist!')
        return -1
    os.chdir(deWork['SHA1'][-9:-1])
    # 提取第一页的地址
    deSoup = bs4.BeautifulSoup(dePage, 'lxml')
    nodeA = deSoup.find('div', attrs={'class': 'gdtm'})
    iniPageUrl = nodeA.find('a').get('href')
    # 循环下载每一页并提取图片的下载链接及下一页的地址
    curPageUrl = ''
    nextPageUrl = iniPageUrl
    pageCnt = 0
    # 循环获取全部
    while nextPageUrl != curPageUrl:
        pageCnt += 1
        print('Deal with page ', pageCnt)
        curPageUrl = nextPageUrl
        curPage = download(url=curPageUrl)
        if curPage is None:
            print('Fail to download page ', pageCnt)
            os.chdir('..')
            rmtree(deWork['SHA1'][-9:-1])
            os.chdir(inPath)
            return 0
        curSoup = bs4.BeautifulSoup(curPage, 'lxml')
        nodeA = curSoup.find('div', attrs={'id': 'i3'})
        curPicUrl = nodeA.find('img').get('src')
        nextPageUrl = nodeA.find('a').get('href')
        # 根据curPicUrl 下载保存图片
        rawPic = download(url=curPicUrl, raw=1, timeout=60)
        if rawPic is None:
            print('**')
            if skip == 1:
                continue
            else:
                os.chdir('..')
                rmtree(deWork['SHA1'][-9:-1])
                os.chdir(inPath)
                return 0
        picFile = open(os.path.basename(curPicUrl), 'wb')
        for chunk in rawPic.iter_content(100000):
            picFile.write(chunk)
        picFile.close()
        # 页间延迟
        time.sleep(randint(1, 3))
    # 下载完成， 重置工作目录
    print('Finish!')
    print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))
    os.chdir(inPath)
    return 1


# 批量下载函数，根据dicFile 中的newDic 下载
# 返回失败的work的ID列表
def downWorkS():
    # 记录入口工作目录
    inPath = os.getcwd()
    # 打开newDic
    dicFile = shelve.open('dicFile')
    newDic = dicFile['newDic']
    dicFile.close()
    # 创建mainFolder保存works
    try:
        os.makedirs('mainFolder')
    except FileExistsError:
        print('MainFolder already exist!')
    os.chdir('mainFolder')
    # 再以当前时间戳创建文件夹
    os.makedirs(str(int(time.time())))
    os.chdir(str(int(time.time())))
    mainPath = os.getcwd()
    # 循环处理每一件work
    workCnt = 0
    folderCnt = 1
    failCnt = []
    for work in newDic.values():
        # 先测试网络
        res = download(url='https://e-hentai.org')
        if res is None:
            print('Fail to connect to e-hentai.org!')
            failCnt.append(work['SHA1'][-9:-1])
            continue
        # 25为单位创建子文件夹
        if workCnt % 25 == 0:
            os.chdir(mainPath)
            os.makedirs(str(folderCnt))
            os.chdir(str(folderCnt))
            folderCnt += 1
        workCnt += 1
        # 提示信息头
        print('Deal with: ', work['title'])
        print(work['SHA1'][-9:-1]+'\t', workCnt, 'of', len(newDic))
        print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))
        # 下载该作品
        res = downWork(work)
        if res == 0:
            failCnt.append(work['SHA1'][-9:-1])
        # 分割线
        print('--------------------------------------------' +
              '--------------------------------------------' +
              '--------------------------------------------')
        # 延时
        time.sleep(randint(5, 10))
    # 全部下载完毕， 提示信息
    print('All work finish!')
    print('Faild work:')
    pprint(failCnt)
    print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))
    # 重置工作区
    os.chdir(inPath)
    # 返回failCnt
    return failCnt


# 创建统计查询表格
# 根据workDic
def makExcel(workDic):
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    sheet.title = 'EH-LOOKUP'
    # 写第一行 属性标签
    sheet.cell(row=1, column=1).value = 'ID'
    sheet.cell(row=1, column=2).value = 'Title'
    sheet.cell(row=1, column=3).value = 'SHA1'
    # 依次写入
    cnt = 1
    for work in workDic.values():
        sheet.cell(row=1+cnt, column=1).value = work['SHA1'][-9:-1]
        sheet.cell(row=1+cnt, column=2).value = work['title']
        sheet.cell(row=1+cnt, column=3).value = work['SHA1']
        cnt += 1
    sheet.freeze_panes = 'B1'
    timeSt = str(int(time.time()))
    wb.save('LOOKUP'+timeSt+'.xlsx')
    print('List saved to LOOKUP.xlsx!')


# 搜索下载
# 根据搜索关键词下载
# 注意： 调用该函数会在当前目录建立新的工作目录 workPath
# 所有文件均保存在新建的工作目录下
def serchDown():
    # 网路检查
    print('Checking the internet connection...')
    res = download(url='https://e-hentai.org')
    if res is None:
        print('Fail to connect to e-hentai.org!')
        exit()
    # 获取搜索参数
    print('设置搜索参数：')
    params = getParam()
    # 获取起止页
    startPage = int(input('起始页（从0开始）：'))
    endPage = int(input('终止页:'))
    while startPage > endPage or startPage < 0 or endPage < 0:
        print('输入错误！')
        print('请重新输入!')
        startPage = int(input('起始页:'))
        endPage = int(input('终止页:'))
    # 根据搜索参数创建主文件夹
    global workPath
    cwPath = os.getcwd()
    workPath = str(params['f_cats']) + str(params['f_search'])
    workPath = os.path.join(cwPath, workPath)
    try:
        os.makedirs(workPath)
    except FileExistsError:
        print('Folder already exists!')
    os.chdir(workPath)
    # 分割线
    print('--------------------------------------------' +
          '--------------------------------------------' +
          '--------------------------------------------')
    # 下载并保存范围内页面
    print('下载搜索结果页:')
    mainUrl = 'https://e-hentai.org'
    htmlFolder = getSearch(url=mainUrl, params=params,
                           startPage=startPage, endPage=endPage)
    # 分割线
    print('--------------------------------------------' +
          '--------------------------------------------' +
          '--------------------------------------------')
    # 从搜索页html中提取每一件work的入口url, 保存在字典中
    print('提取入口url:')
    getWorkUrl(htmlFolder)
    # 分割线
    print('--------------------------------------------' +
          '--------------------------------------------' +
          '--------------------------------------------')
    # 批量下载所有的work 并删除失败项
    print('批量下载：')
    failList = downWorkS()
    if len(failList) > 0:
        dicFile = shelve.open('dicFile')
        workDic = dicFile['workDic']
        for workID in failList:
            del workDic[workID]
        dicFile['workDic'] = workDic
        dicFile.close()
        del workDic
    # 根据dicFile创建查询表
    dicFile = shelve.open('dicFile')
    workDic = dicFile['workDic']
    dicFile.close()
    makExcel(workDic)
    # 删除搜索页
    rmtree('scHtml')


# 通过给定详情页url 直接下载
# 成功 返回deWork 字典
# 无法下载详情页 返回-1
# 无法提取title 返回-2
# 下载失败或缺页 返回-3
def downDirect(deUrl):
    # 下载详情页(测试网络)
    dePage = download(url=deUrl)
    if dePage is None:
        print('Fail to download detail page!')
        return -1
    # 提取title 并计算SHA1 构造deWork 字典
    deWork = {}
    deSoup = bs4.BeautifulSoup(dePage, 'lxml')
    title = deSoup.find('h1', attrs={'id': 'gn'}).getText()
    if len(title) == 0:
        print('Fial to extract title!')
        return -2
    deWork['title'] = title
    workSHA1 = sha1(deWork['title'].encode(
        encoding='UTF-8')).hexdigest()
    deWork['SHA1'] = workSHA1
    deWork['url'] = deUrl
    # 根据deWork 下载work
    print('Deal with: '+title)
    res = downWork(deWork)
    if res != 1:
        print('Something thing wrong happend while downloading!')
        return -3
    else:
        print('Download successful!')
        return deWork


# 从excel中获取详情页url 直接下载
def excelDown(excelPath):
    # 记录入口工作目录
    inPath = os.getcwd()
    # 打开表格
    try:
        wb = openpyxl.load_workbook(excelPath)
    except FileNotFoundError:
        print('表格打开失败！')
        return -1
    sheet = wb.get_active_sheet()
    # 获取url 批量下载 记录在workDic中
    i = 0
    workDic = {}
    failWork = []
    deUrl = sheet.cell(row=1+i, column=1).value
    while deUrl is not None:
        i += 1
        print('Work count:', i)
        tempDic = downDirect(deUrl)
        if tempDic == -1 or tempDic == -2 or tempDic == -3:
            failWork.append(deUrl)
            continue
        workDic[tempDic['SHA1'][-9:-1]] = copy.copy(tempDic)
        deUrl = sheet.cell(row=1+i, column=1).value
        # 分割线
        print('--------------------------------------------' +
              '--------------------------------------------' +
              '--------------------------------------------')
        # 延时
        time.sleep(randint(1, 5))
    # 制作查询表
    if len(workDic) > 0:
        makExcel(workDic)
        print('All work finish!')
        if len(failWork) > 0:
            print('Faild work:')
            pprint(failWork)
        print('Look up list saved to LOOKUP.xlsx')
    # 恢复工作目录
    os.chdir(inPath)


# 测试
if __name__ == '__main__':
    print('搜索下载 or 直接下载？')
    choice = input('1.搜索下载  2.直接下载  3.从表格下载\n')
    if choice == '1':
        serchDown()
    elif choice == '2':
        deUrl = input('请务必输入正确的详情页网址：\n')
        downDirect(deUrl, skip=1)
    elif choice == '3':
        excelPath = input('输入表格路径：\n')
        excelDown(excelPath)
    else:
        print('Invalid input!')
        exit()
