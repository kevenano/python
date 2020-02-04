# EOP 批量采集
import requests
import bs4
import os
import time
import shelve
import re
from random import randint
from pprint import pprint
import copy
import openpyxl

Cookie = 'menunew=0-0-0-0; ' +\
    'username=kevenano; ' +\
    'password=qhm2012%40%40%40; ' + \
    'remember=1; ' +\
    'huiyuan=' +\
    '7f39Vfi514lH9dndqdzF3wpjjvZy1TiC9eReu%2FmI2RXQTumC1OpKxh7MrBE; ' +\
    'FWE_getuser=kevenano; ' +\
    'FWE_getuserid=161673; ' +\
    'login_mima=9161d8bd2056459a0a96f80f2e6a818d; ' +\
    'PHPSESSID=75e2ed7a61e14459313010386cb2ecdc; ' +\
    'think_language=en-US'


# Html download function
def download(url, num_retries=3, cookie='', raw=0, timeout=40):
    print('Downloading: ', url)
    headers = {
        'user-agent':
        'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:72.0)' +
        ' Gecko/20100101 Firefox/72.0',
        'connection':
        'keep-alive'
    }
    if cookie != '':
        headers['cookie'] = cookie
    try:
        resp = requests.get(url, headers=headers, timeout=timeout)
        html = resp.text
        if resp.status_code >= 400:
            print('Download error: ', resp.text)
            html = None
            if num_retries and 500 <= resp.status_code < 600:
                return download(url, num_retries-1)
    except requests.exceptions.RequestException:
        print('Download error!!!')
        html = None
        resp = None
    except requests.exceptions.Timeout:
        print('请求超时!')
        html = None
        resp = None
    if raw == 1:
        return resp
    else:
        return html


# 获取关键头 即 work的ID&Title字段
# 保存在keyFile中
# 输入save_html=1 保存页面
# 输入classUrl 对应类别曲谱分类首页
# 如：https://www.everyonepiano.cn/Music-class12-动漫.html
# 关键是为了获得class12 和动漫 两个字段
# canshu: 按热门排序'clicks' 按最新发布顺序'cn_edittime'
# pageAll 按排列顺序需要下载的总页数
def getKey(classUrl, pageStart, pageEnd, canshu='clicks', save_html=1):
    pageCnt = pageStart
    keys = []
    try:
        os.makedirs('Html')
    except FileExistsError:
        print('Folder already exist!')
    while pageCnt <= pageEnd:
        # Download & Save Html File
        print('Deal with page ', pageCnt)
        pageUrl = classUrl + '?come=web&p=' +\
            str(pageCnt) +\
            '&canshu=' + canshu +\
            '&word=&author=&jianpu=&paixu=desc&username='
        pageHtml = download(url=pageUrl, cookie=Cookie)
        if pageHtml is None:
            continue
        if bool(save_html):
            htmlFile = open('Html/listPage'+str(pageCnt)+'.html', 'w')
            htmlFile.write(pageHtml)
            htmlFile.close()
        # 提取Key
        pageSoup = bs4.BeautifulSoup(pageHtml, 'lxml')
        allElem = pageSoup.find_all('a', attrs={'class': 'Title'})
        if len(allElem) < 2:
            continue
        for i in range(1, len(allElem)):
            keys.append(allElem[i].get('href').replace('/Music-', ''))
        # 计数+1
        pageCnt += 1
        # 延时
        print('Sleeping...')
        time.sleep(randint(1, 5))

    # 保存keys变量
    keyFile = shelve.open('keyFile')
    keyFile['keys'] = keys
    keyFile.close()
    print('Keys saved to keyFile!')


# 构造资源下载页URL字典
# 字典保存在 urlFile 中 workDic 字典变量中
#  以id 为索引
# 包括: pdf midi eopn title id
def buildUrl():
    # 获取keys
    keyFile = shelve.open('keyFile')
    keys = keyFile['keys']
    # 构造URL 字典
    workDic = {}
    urlDic = {}
    titleReg = re.compile(r'\d+-(.+)\.html')
    idReg = re.compile(r'(\d+)')
    for urlKey in keys:
        pdfPage = 'https://www.everyonepiano.cn/PDF-'+urlKey
        midiPage = 'https://www.everyonepiano.cn/Midi-'+urlKey
        eopnPage = 'https://www.everyonepiano.cn/Eopn-down-'+urlKey
        eopmPage = 'https://www.everyonepiano.cn/Eopm-down-'+urlKey
        title = titleReg.search(urlKey).group(1)
        workId = idReg.search(urlKey).group(1)
        urlDic['pdf'] = pdfPage
        urlDic['midi'] = midiPage
        urlDic['eopn'] = eopnPage
        urlDic['eopm'] = eopmPage
        urlDic['title'] = title
        urlDic['id'] = workId
        workDic[workId] = copy.copy(urlDic)  # 小心！！！！！！！
    # 保存字典
    urlFile = shelve.open('urlFile')
    urlFile['workDic'] = workDic
    urlFile.close()
    print('Page urls saved to urlFile!')


# 改进版 下载函数
# 输入： workDic 包含下载页面链接的字典 colNum 当前处理卷的卷数
# 返回： downDic 包含直接资源下载链接的字典 failCnt
def downCol(workDic, colNum):
    # 记录输入工作目录
    inDir = os.getcwd()
    # 计数用
    count = 1
    # 获取下载url 保存在downUrl
    downUrl = {}
    # 用于记录下载失败的项
    failCnt = {}
    for work in workDic.values():
        # 提示信息头
        print('Deal with '+work['title'])
        print(count, 'of', len(workDic), 'in col', colNum)
        count += 1
        tempCnt = []

        # 在当前目录下创建该work专属文件夹
        # 若已存在 则跳过
        try:
            os.makedirs(work['id']+'_'+work['title'])
        except FileExistsError:
            print('File already exist!')
            continue
        # 将工作区移入卷文件夹
        os.chdir(work['id']+'_'+work['title'])
        # 用于暂存的字典
        # 包括 pdfS  pdfN midi eopn
        #     五线谱 简谱
        itemUrl = {}

        # pdf down
        # 创建pdf文件夹 保存pdf相关资源
        os.makedirs('pdf')
        pdfHtml = download(url=work['pdf'], cookie=Cookie)
        if pdfHtml is None:
            # 若网页为空 设置对应下载链接为空
            itemUrl['pdfS'] = ''
            itemUrl['pdfN'] = ''
            tempCnt.append('pdf')
        else:
            # 保存网页
            htmlFile = open('pdf/'+work['id']+'.html', 'w')
            htmlFile.write(pdfHtml)
            htmlFile.close()
            # 提取下载链接
            pdfSoup = bs4.BeautifulSoup(pdfHtml, 'lxml')
            urlElem = pdfSoup.find_all('a', attrs={'class': 'btn-success'})
            if len(urlElem) < 2:
                # 找不到链接 设置对应下载链接为空
                itemUrl['pdfS'] = ''
                itemUrl['pdfN'] = ''
                tempCnt.append('pdf')
            else:
                pdfS = 'https://www.everyonepiano.cn' +\
                    urlElem[0].get('href')
                pdfN = 'https://www.everyonepiano.cn' +\
                    urlElem[1].get('href')
                # 将链接存入字典
                itemUrl['pdfS'] = pdfS
                itemUrl['pdfN'] = pdfN
                # 直接下载
                # 五线谱
                pdfSRes = download(url=pdfS, cookie=Cookie, raw=1)
                if pdfSRes is None:
                    tempCnt.append('pdfS')
                else:
                    pdfSFile = open('pdf/'+work['title']+'-五线谱'+'.pdf', 'wb')
                    for chunk in pdfSRes.iter_content(100000):
                        pdfSFile.write(chunk)
                    pdfSFile.close()
                # 简谱
                pdfNRes = download(url=pdfN, cookie=Cookie, raw=1)
                if pdfNRes is None:
                    tempCnt.append('pdfN')
                else:
                    pdfNFile = open('pdf/'+work['title']+'-简谱'+'.pdf', 'wb')
                    for chunk in pdfNRes.iter_content(100000):
                        pdfNFile.write(chunk)
                    pdfNFile.close()

        # midi down
        # 创建midi文件夹 保存midi资源
        os.makedirs('midi')
        midiHtml = download(url=work['midi'], cookie=Cookie)
        if midiHtml is None:
            itemUrl['midi'] = ''
            tempCnt.append('midi')
        else:
            # 保存html
            htmlFile = open('midi/'+work['id']+'.html', 'w')
            htmlFile.write(midiHtml)
            htmlFile.close()
            # 提取下载链接
            midiSoup = bs4.BeautifulSoup(midiHtml, 'lxml')
            urlElem = midiSoup.find_all('a', attrs={'class': 'btn-success'})
            if len(urlElem) < 1 or urlElem[0].get('href') == '#':
                # 未找到下载链接
                itemUrl['midi'] = ''
                tempCnt.append('midi')
            else:
                midi = 'https://www.everyonepiano.cn' +\
                    urlElem[0].get('href')
                # 写入字典
                itemUrl['midi'] = midi
                # 直接下载
                midiRes = download(url=midi, cookie=Cookie, raw=1)
                if midiRes is None:
                    tempCnt.append('midi')
                else:
                    midiFile = open('midi/'+work['title']+'.mid', 'wb')
                    for chunk in midiRes.iter_content(100000):
                        midiFile.write(chunk)
                    midiFile.close()

        # eopn down
        # 创建eopn文件夹 保存eopn资源
        os.makedirs('eopn')
        eopnHtml = download(url=work['eopn'], cookie=Cookie)
        if eopnHtml is None:
            itemUrl['eopn'] = ''
            tempCnt.append('eopn')
        else:
            # 保存html
            htmlFile = open('eopn/'+work['id']+'.html', 'w')
            htmlFile.write(eopnHtml)
            htmlFile.close()
            # 提取下载链接
            eopnSoup = bs4.BeautifulSoup(eopnHtml, 'lxml')
            urlElem = eopnSoup.find_all('a', attrs={'class': 'btn-success'})
            if len(urlElem) < 1 or urlElem[0].get('href') == '#':
                itemUrl['eopn'] = ''
                tempCnt.append('eopn')
            else:
                eopn = 'https://www.everyonepiano.cn' +\
                    urlElem[0].get('href')
                # 写入字典
                itemUrl['eopn'] = eopn
                # 直接下载
                eopnRes = download(url=eopn, cookie=Cookie, raw=1)
                if eopnRes is None:
                    tempCnt.append('eopn')
                else:
                    eopnFile = open('eopn/'+work['title']+'.eopn', 'wb')
                    for chunk in eopnRes.iter_content(100000):
                        eopnFile.write(chunk)
                    eopnFile.close()

        # 将当前资源的所有下载链接写入字典 downUrl
        downUrl[work['id']] = itemUrl
        # 如果有失败项 写入字典
        if len(tempCnt) > 0:
            failCnt[work['id']] = copy.copy(tempCnt)
        # 将工作区移至父文件夹
        os.chdir('..')
        # 休息3~10秒
        print('Sleeping...')
        print(time.asctime(time.localtime(time.time())))
        time.sleep(randint(3, 10))
    # 返回包含下载链接字典 以及失败项
    # 同时重置工作目录
    os.chdir(inDir)
    return downUrl, failCnt


# 创建统计查询表格
# 根据workDic
def makExcel(workDic):
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    sheet.title = 'EOP_ANIME'
    # 写第一行 属性标签
    sheet.cell(row=1, column=1).value = 'Vol'
    sheet.cell(row=1, column=2).value = 'ID'
    sheet.cell(row=1, column=3).value = 'Tiltle'
    # 依次写入
    cnt = 1
    vol = 1
    for work in workDic.values():
        sheet.cell(row=1+cnt, column=1).value = vol
        sheet.cell(row=1+cnt, column=2).value = work['id']
        sheet.cell(row=1+cnt, column=3).value = work['title']
        if cnt % 30 == 0:
            vol += 1
        cnt += 1
    sheet.freeze_panes = 'B1'
    wb.save('eopList.xlsx')
    print('Data saved to eopList.xlsx!')


# 下载 & 本地保存策略
# 先以30为单位分卷
# 输入可选参数 起始卷 终止卷（包含） 卷列表
def saveDown(classChoos='mainFolder', startCol=1, endCol=999, sepCol=[]):
    # 记录入口工作目录
    inPath = os.getcwd()
    # 打开work字典
    urlFiel = shelve.open('urlFile')
    workDic = urlFiel['workDic']
    urlFiel.close()
    # 分割字典
    workList = []
    tempDic = {}
    count = 0
    for work in workDic.values():
        tempDic[work['id']] = copy.copy(work)
        count += 1
        if count % 30 == 0:
            # 每30个为一卷
            workList.append(tempDic)
            tempDic = {}
    if len(tempDic) != 0:
        workList.append(tempDic)
    print('字典分割完毕，共计 ', len(workList), '卷', ' ', len(workDic), '项')
    print('准备开始处理...')
    time.sleep(10)
    print('开始处理:')
    print(time.asctime(time.localtime(time.time())))
    # 记录处理了的卷
    colList = []
    # 下载卷数处理
    if startCol < 1:
        startCol = 1
    if endCol > len(workList):
        endCol = len(workList)
    # 创建主文件夹
    try:
        os.makedirs(classChoos)
    except FileExistsError as FEE:
        print(str(FEE))
    # 分两种策略处理
    if len(sepCol) == 0:
        for col in range(startCol-1, endCol):
            print('正在处理第 ', col+1, '卷', '...')
            # 重置主工作区
            os.chdir(workPath+'/'+classChoos)
            # 创建卷文件夹
            try:
                os.makedirs(str(col+1))
            except FileExistsError:
                print('Col ', col+1, ' already here!')
                '''
                # 若卷文件夹下包含 srcFile
                # 说明该卷已处理完毕
                # 跳过改卷
                # 否则继续处理该卷
                if 'srcFile' in os.listdir(str(col+1)):
                    continue
                '''
            # 移动工作区至卷文件夹内
            os.chdir(str(col+1))
            # 处理卷
            downUrl, failCnt = downCol(workList[col], col+1)
            colList.append(col+1)
            print('第 ', col+1, '卷', '处理完毕！')
            # 处理失败项
            if len(failCnt) > 0:
                print('失败项：')
                pprint(failCnt)
                # 保存失败项
                failFile = shelve.open('failFile')
                failFile['failCnt'] = failCnt
                failFile.close()
                print('失败项保存在failFile中')
            # 保存下载链接
            if 'srcFile' not in os.listdir():
                srcFile = shelve.open('srcFile')
                srcFile['downUrl'] = downUrl
                srcFile.close()
                print('下载链接保存在srcFile中')
                print('沉默1~2分钟...')
                time.sleep(randint(60, 120))
    else:
        for col in sepCol:
            print('正在处理第 ', col, '卷', '...')
            if col < 1 or col > len(workList):
                print('卷标超出范围!')
                continue
            # 重置主工作区
            os.chdir(workPath+'/'+classChoos)
            # 创建卷文件夹
            try:
                os.makedirs(str(col))
            except FileExistsError:
                print('Col ', col, ' already here!')
                '''
                # 若卷文件夹下包含 srcFile
                # 说明该卷已处理完毕
                # 跳过改卷
                # 否则继续处理该卷
                if 'srcFile' in os.listdir(str(col)):
                    continue
                '''
            # 移动工作区至卷文件夹内
            os.chdir(str(col))
            # 处理卷
            downUrl, failCnt = downCol(workList[col-1], col)
            colList.append(col)
            print('第 ', col, '卷', '处理完毕！')
            # 处理失败项
            if len(failCnt) > 0:
                print('失败项：')
                pprint(failCnt)
                # 保存失败项
                failFile = shelve.open('failFile')
                failFile['failCnt'] = failCnt
                failFile.close()
                print('失败项保存在failFile中')
            # 保存下载链接
            if 'srcFile' not in os.listdir():
                srcFile = shelve.open('srcFile')
                srcFile['downUrl'] = downUrl
                srcFile.close()
                print('下载链接保存在srcFile中')
                print('沉默1~2分钟...')
                time.sleep(randint(60, 120))
    print('任务完成！')
    print('总共处理 ', len(colList), '卷')
    print('本次处理的卷:')
    print(colList)
    print(time.asctime(time.localtime(time.time())))
    # 还原路径
    os.chdir(inPath)


# 全流程
# 以轻音乐为例
def main():
    classUrl = {}
    classUrl['轻音乐'] = 'https://www.everyonepiano.cn/Music-class16-轻音乐.html'
    classUrl['流行'] = 'https://www.everyonepiano.cn/Music-class5-流行.html'
    classUrl['影视'] = 'https://www.everyonepiano.cn/Music-class31-影视.html'
    classUrl['经典'] = 'https://www.everyonepiano.cn/Music-class11-经典.html'
    classUrl['动漫'] = 'https://www.everyonepiano.cn/Music-class12-动漫.html'
    # 获取处理类
    classChoos = input('输入曲谱类别：')
    while classChoos not in classUrl:
        print('没有该类别!')
        classChoos = input('请重新输入：')
        if classChoos == 'exit' or classChoos == 'quit':
            exit()
    # 获取排序标准
    print('选择排序方式：')
    print('0.热门  1. 新发布')
    canshu = input()
    if canshu == '0':
        canshu = 'clicks'
    else:
        canshu = 'cn_edittime'
    # 获取处理页数
    try:
        pageStart = int(input('输入处理起始页:'))
        pageEnd = int(input('输入处理结束页:'))
    except ValueError:
        print('输入错误！')
        print('只处理前10页')
        pageStart = 1
        pageEnd = 10
    # 根据输入信息创建工作目录
    global workPath
    workPath = os.getcwd()+'/'+classChoos+'-'+canshu+'-' +\
        str(pageStart) + 'to'+str(pageEnd)
    try:
        os.makedirs(workPath)
    except FileExistsError:
        print('Folder already exist!')
    os.chdir(workPath)
    # 先检查当前目录是否存在keyFIle
    if 'keyFile' not in os.listdir():
        # 按顺序获取每一页中 每一项的关键头 得到keyFile
        getKey(classUrl=classUrl[classChoos],
               pageStart=pageStart, pageEnd=pageEnd,
               canshu=canshu, save_html=1)
    # 检查urlFile
    if 'urlFile' not in os.listdir():
        # 构造资源下载页URL字典 得到urlFile
        buildUrl()
    # 下载 & 本地保存策略
    saveDown(classChoos=classChoos)
    # 导出查询表
    urlFiel = shelve.open('urlFile')
    workDic = urlFiel['workDic']
    urlFiel.close()
    makExcel(workDic)
    # 结束
    print('END')
    print(time.asctime(time.localtime(time.time())))


# 测试
if __name__ == '__main__':
    main()
