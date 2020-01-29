# DLsite 数据收集
import requests
import bs4
import openpyxl
import os
import time
import shelve
import json
import re
# import logging


os.chdir('/home/kevenano/Music/Works')

# Get work_id from xlsx


def get_id(tableFile, start_row=2, start_column=1):
    wb = openpyxl.load_workbook(tableFile)
    sheet = wb.get_active_sheet()
    work_ids = []
    i = 0
    while bool(sheet.cell(row=start_row+i, column=start_column).value):
        work_ids.append(sheet.cell(row=start_row+i, column=start_column).value)
        i += 1
    return work_ids


# Html download function
def download(url, num_retries=3):
    print('Downloading: ', url)
    headers = {
        'user-agent':
        'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:72.0) ' +
        'Gecko/20100101 Firefox/72.0'
    }
    try:
        resp = requests.get(url, headers=headers)
        html = resp.text
        if resp.status_code >= 400:
            print('Download error: ', resp.text)
            html = None
            if num_retries and 500 <= resp.status_code < 600:
                return download(url, num_retries-1)
    except requests.exceptions.RequestException:
        print('Download error')
        html = None
    return html


def test1():
    # Get work_id
    work_ids = get_id('ASMR_alpha.xlsx')
    # 记录所有出现过的属性
    attrSum = []
    attrSum.append('ID')
    # 循环取得每个work的属性信息，存在works中
    works = {}
    for work_id in work_ids:
        # Download&Save html
        mainHtml = download(
            'https://www.dlsite.com/maniax/work/=/product_id/'+work_id)
        if mainHtml is None:
            continue
        htmlFile = open('/home/kevenano/Music/Works/' +
                        work_id+'.html', 'w')
        htmlFile.write(mainHtml)
        htmlFile.close()

        mainSoup = bs4.BeautifulSoup(mainHtml, features='lxml')
        work_header = bs4.BeautifulSoup(
            str(mainSoup.select('#work_header')[0]), features='lxml')
        work_left = bs4.BeautifulSoup(
            str(work_header.select('#work_left')[0]), features='lxml')
        work_right = bs4.BeautifulSoup(
            str(work_header.select('#work_right')[0]), features='lxml')
        work_value = bs4.BeautifulSoup(
            str(work_left.select('#work_value')[0]), features='lxml')

        # 属性名
        attrName = work_right.find_all('th')
        # 属性值
        attrValue = work_right.find_all('td')

        # 写work字典
        workAttr = {}
        workAttr['ID'] = work_id
        for i in range(len(attrName)):
            key = attrName[i].getText().replace('\n', '')
            value = attrValue[i].getText().replace('\n', '').replace('  ', '')
            workAttr[key] = value
            # 更新出现过的属性类
            if key not in attrSum:
                attrSum.append(key)
        works[work_id] = workAttr

        # 延迟
        print('Sleeping...')
        time.sleep(5)

    # 写表格
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    sheet.title = 'DL works V1'
    # 写第一行 属性标签
    for i in range(len(attrSum)):
        sheet.cell(row=1, column=1+i).value = attrSum[i]
    # 依次写入每件work的属性
    i = 0
    for ID, attrs in works.items():
        for attr, attrV in attrs.items():
            sheet.cell(row=i+2, column=attrSum.index(attr)+1).value = attrV
        i += 1

    # 表格格式处理
    # 冻结第一列
    sheet.freeze_panes = 'B1'

    # 保存表格文件
    wb.save('output_gama.xlsx')
    print('Data saved to '+'output_gama.xlsx')


# 离线
def test2():
    # Get work_id
    work_ids = get_id('ASMR_alpha.xlsx')
    # 记录所有出现过的属性
    attrSum = []
    attrSum.append('ID')
    # 循环取得每个work的属性信息，存在works中
    works = {}
    for work_id in work_ids:
        try:
            print('Deal with '+work_id+'...')
            mainHtml = open(work_id+'.html', 'r')
        except FileNotFoundError as e:
            print(str(e))
            continue

        mainSoup = bs4.BeautifulSoup(mainHtml, features='lxml')
        work_header = bs4.BeautifulSoup(
            str(mainSoup.select('#work_header')[0]), features='lxml')
        work_left = bs4.BeautifulSoup(
            str(work_header.select('#work_left')[0]), features='lxml')
        work_right = bs4.BeautifulSoup(
            str(work_header.select('#work_right')[0]), features='lxml')
        work_value = bs4.BeautifulSoup(
            str(work_left.select('#work_value')[0]), features='lxml')

        # 属性名
        attrName = work_right.find_all('th')
        # 属性值
        attrValue = work_right.find_all('td')

        # 写work字典
        workAttr = {}
        workAttr['ID'] = work_id
        for i in range(len(attrName)):
            key = attrName[i].getText().replace('\n', '')
            value = attrValue[i].getText().replace('\n', '').replace('  ', '')
            workAttr[key] = value
            # 更新出现过的属性类
            if key not in attrSum:
                attrSum.append(key)
        works[work_id] = workAttr

    # 写表格
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    sheet.title = 'DL works V1'
    # 写第一行 属性标签
    for i in range(len(attrSum)):
        sheet.cell(row=1, column=1+i).value = attrSum[i]
    # 依次写入每件work的属性
    i = 0
    for ID, attrs in works.items():
        for attr, attrV in attrs.items():
            sheet.cell(row=i+2, column=attrSum.index(attr)+1).value = attrV
        i += 1

    # 表格格式处理
    # 冻结第一列
    sheet.freeze_panes = 'B1'

    # 保存表格文件
    wb.save('output_beta.xlsx')
    print('Data saved to '+'output_beta.xlsx')


# 直接爬取work_value相关数据
def test3():
    # Get work_id
    work_ids = get_id('ASMR_alpha.xlsx')
    # 选择所需属性
    attrNeed = ['dl_count', 'rate_average_2dp',
                'rate_count', 'review_count', 'wishlist_count']
    # 循环取得每个work的信息，存在works中
    works = {}
    for work_id in work_ids:
        # 下载&保存数据字典 原始字符串保存在dicStr中
        resUrl = 'https://www.dlsite.com/maniax/product/info/ajax?product_id=' + \
            work_id+'&cdn_cache_min=1'
        dicStr = download(resUrl)
        if dicStr is None:
            continue
        dicFile = shelve.open(work_id)
        dicFile['dicStr'] = dicStr
        dicFile.close()
        # 将字符串转换为字典变量
        mainDic = json.loads(dicStr)
        attrDic = mainDic[work_id]
        # 写works字典
        workAttr = {}
        workAttr['ID'] = work_id
        for attr in attrNeed:
            if attr not in attrDic:
                workAttr[attr] = ''
            else:
                workAttr[attr] = attrDic[attr]
        works[work_id] = workAttr
        # 延迟
        print('Sleeping...')
        time.sleep(5)
    # 写表格
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    sheet.title = 'DL works V2'
    # 写第一行 属性标签
    sheet.cell(row=1, column=1).value = 'ID'
    for i in range(len(attrNeed)):
        sheet.cell(row=1, column=2+i).value = attrNeed[i]
    # 依次写入所有属性数据
    i = 0
    for ID, attrs in works.items():
        sheet.cell(row=i+2, column=1).value = ID
        for attr in attrNeed:
            sheet.cell(row=i+2, column=attrNeed.index(attr) +
                       2).value = attrs[attr]
        i += 1
    # 表格格式处理
    # 冻结第一列
    sheet.freeze_panes = 'B1'
    # 保存表格文件
    wb.save('output_sigma.xlsx')
    print('Data saved to '+'output_sigma.xlsx')


# 离线
def test4():
    # Get work_id
    work_ids = get_id('ASMR_alpha.xlsx')
    # 选择所需属性
    attrNeed = ['dl_count', 'rate_average_2dp',
                'rate_count', 'review_count', 'wishlist_count']
    # 循环取得每个work的信息，存在works中
    works = {}
    for work_id in work_ids:
        # 打开work对应字典数据
        try:
            print('Deal with '+work_id+'...')
            dicStr = shelve.open(work_id)['dicStr']
        except FileNotFoundError as e:
            print(str(e))
            continue
        # 将字符串转换为字典变量
        mainDic = json.loads(dicStr)
        attrDic = mainDic[work_id]
        # 写works字典
        workAttr = {}
        workAttr['ID'] = work_id
        for attr in attrNeed:
            if attr not in attrDic:
                workAttr[attr] = ''
            else:
                workAttr[attr] = attrDic[attr]
        works[work_id] = workAttr
    # 写表格
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    sheet.title = 'DL works V2'
    # 写第一行 属性标签
    sheet.cell(row=1, column=1).value = 'ID'
    for i in range(len(attrNeed)):
        sheet.cell(row=1, column=2+i).value = attrNeed[i]
    # 依次写入所有属性数据
    i = 0
    for ID, attrs in works.items():
        sheet.cell(row=i+2, column=1).value = ID
        for attr in attrNeed:
            sheet.cell(row=i+2, column=attrNeed.index(attr) +
                       2).value = attrs[attr]
        i += 1
    # 表格格式处理
    # 冻结第一列
    sheet.freeze_panes = 'B1'
    # 保存表格文件
    wb.save('output_sigma.xlsx')
    print('Data saved to '+'output_sigma.xlsx')


# 获取封面图
def test5():
    # Get work_id
    work_ids = get_id('ASMR_alpha.xlsx')
    for work_id in work_ids:
        # Download&Save html
        mainHtml = download(
            'https://www.dlsite.com/maniax/work/=/product_id/'+work_id)
        if mainHtml is None:
            continue
        htmlFile = open('/home/kevenano/Music/Works/' +
                        work_id+'.html', 'w')
        htmlFile.write(mainHtml)
        htmlFile.close()
        # bs4
        mainSoup = bs4.BeautifulSoup(mainHtml, features='lxml')
        coverElem = mainSoup.select('.slider_items trans')
        if coverElem == []:
            print('Could not find cover image.')
            continue
        else:
            coverUrl = coverElem[0].get('src')
            # Download the image
            print('Downloading image '+work_id)
            res = requests.get(coverUrl)
            res.raise_for_status()
            # Save the image
            imageFile = open(os.path.basename(coverUrl), 'wb')
            for chunk in res.iter_content(100000):
                imageFile.write(chunk)
            imageFile.close()
        # 延时
        print('Sleeping...')
        time.sleep(5)


# 离线
def test6():
    # Get work_id
    work_ids = get_id('ASMR_alpha.xlsx')
    for work_id in work_ids:
        try:
            print('Deal with '+work_id+':')
            mainHtml = open(work_id+'.html', 'r')
        except FileNotFoundError as e:
            print(str(e))
            continue
        # bs4
        mainSoup = bs4.BeautifulSoup(mainHtml, features='lxml')
        coverStr = str(mainSoup.select('.product-slider-data'))
        coverRegex = re.compile('src="([^"]{1,})"')
        coverElem = coverRegex.findall(coverStr)
        if coverElem == []:
            print('Could not find cover image.')
            continue
        else:
            for elem in coverElem:
                coverUrl = 'https:'+elem
                # Download the image
                print('Downloading '+os.path.basename(coverUrl)+'...')
                res = requests.get(coverUrl)
                res.raise_for_status()
                # Save the image
                imageFile = open(os.path.basename(coverUrl), 'wb')
                for chunk in res.iter_content(100000):
                    imageFile.write(chunk)
            imageFile.close()
        # 延时
        print('Sleeping...')
        time.sleep(5)
    print('Cover download finiah!')


if __name__ == '__main__':
    test6()
